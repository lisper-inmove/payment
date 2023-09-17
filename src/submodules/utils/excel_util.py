# -*- coding: utf-8 -*-

import xlsxwriter
import openpyxl


class ExcelUtil:

    def load(self, stream, skipline=None):
        wb = openpyxl.load_workbook(stream)
        return self.__read_from_wb(wb, skipline=skipline)

    def read(self, filepath, read_only=False, skipline=None):
        wb = openpyxl.open(filepath, read_only=read_only)
        return self.__read_from_wb(wb, skipline=skipline)

    def __read_from_wb(self, wb, skipline=None):
        if skipline is None:
            skipline = 0
        result = []
        for sheet in wb:
            for row in sheet.iter_rows():
                if skipline > 0:
                    skipline -= 1
                    continue
                cell_values = []
                for cell in row:
                    cell_values.append(cell.value)
                result.append(cell_values)
        return result

    def write(self, filepath, values):
        wb = xlsxwriter.Workbook(filepath)
        ws = wb.add_worksheet()
        row = 0
        for index, value in enumerate(values):
            ws.write(row, index, value)
        row += 1
        wb.close()


if __name__ == '__main__':
    print(ExcelUtil().read("/home/inmove/Desktop/kstest.xlsx"))
